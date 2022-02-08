from openpyxl import load_workbook
from datetime import *

##Class for represnting a type of PO
##Created to allow easy attribute access, hashing, and comparisons
class PO:

    def __init__(self):
        self.ESD = None
        self.ESDDate = None
        self.Supplier = None
        self.Origin = None
        self.Destination = None
        self.Equipment = None
        self.FEU = None

    def __hash__(self):
        return hash((self.ESD, self.Supplier, self.Origin, self.Destination, self.Equipment, self.FEU))

    def __str__(self):
        return self.Equipment + " container of size " + str(self.FEU) + " sent by " + self.Supplier + " from " + self.Origin + " to " + self.Destination + " on " + self.ESDDate.strftime("%m/%d/%y")

    def __eq__(self, other):
        return self.ESD == other.ESD and self.Supplier == other.Supplier and self.Origin == other.Origin and self.Destination == other.Destination and self.Equipment == other.Equipment and self.FEU == other.FEU and self.ESDDate == other.ESDDate

##Class for representing assignments
class Assignment:

    def __init__(self):
        self.Carrier = None
        self.Origin = None
        self.Destination = None
        self.CYClose = None
        self.ETD = None
        self.TTime = None
        self.String = None
        self.Vessel = "Unknown"

    def similar(self, other):
        return self.Carrier == other.Carrier and HomeDepotData.cityEqual(self.Origin, other.Origin) and HomeDepotData.cityEqual(self.Destination, other.Destination) and self.ETD == other.ETD and self.TTime == other.TTime and self.String == other.String

    def __hash__(self):
        return hash((self.Carrier, self.Origin, self.Destination, self.CYClose, self.ETD, self.TTime, self.String))

    def __str__(self):
        return self.Carrier + " traveling from " + self.Origin + " to " + self.Destination + " on string " + str(self.String) + " on day " + str(self.ETD) + " with TTime " + str(self.TTime)

    def __eq__(self, other):
        return self.Carrier == other.Carrier and self.Origin == other.Origin and self.Destination == other.Destination and self.CYClose == other.CYClose and self.ETD == other.ETD and self.TTime == other.TTime and self.String == other.String


##Class that holds all problem data to be inserted into the model
##Created so that all data can be placed into a single object and plugged into the model
class HomeDepotData:

    #For "Equipment" field
    #Converts from the format in the PO data files to the format in the schedule data file
    #If an equipment designation cannot be found in the list below, the equipment name is left unchanged
    equipmentConversionDict = {"40 FT. HIGH CUBE CONTAINER" : "40HIGH",
                               "20 FT. CONTAINER" : "20DRY",
                               "40 FT. HIGH CUBE REFRIGERATED CONT." : "40' High Cube Non Operating Reefer",
                               "45 FT. CONTAINER" : "45HIGH",
                               "40 FT. STANDARD CONTAINER" : "40DRY",
                               "CFS: LESS THAN CONTAINER LOAD" : "CFS",
                               "40HREF" : "40' High Cube Non Operating Reefer",
                               "40REEF" : "40' High Cube Non Operating Reefer"}
    @staticmethod
    def equipmentConversion(equip):
        equip = str(equip)
        if equip in HomeDepotData.equipmentConversionDict: return HomeDepotData.equipmentConversionDict[equip]
        else: return equip

    #For "FEU" field
    #This only exists in the case that the data doesn't have an FEU column
    feuConversionDict = {"40HIGH" : 1.125,
                         "20DRY" : .5,
                         "40' High Cube Non Operating Reefer" : 1,
                         "45HIGH" : 1.25,
                         "40DRY" : 1,
                         "CFS" : .26}
    @staticmethod
    def feuConversion(equip):
        equip = str(equip)
        if equip in HomeDepotData.feuConversionDict: return HomeDepotData.feuConversionDict[equip]
        else: return 1
                    


    #For "Origin" field
    #Converts from the format in the PO data files to the format in the schedule data file
    #If a city cannot be found in the list below, the city name is extracted and changed to upper case
    #Example: "Dallas, TX" would be changed to "DALLAS"
    originConversionDict = {"CHEFOO/YANTAI/YENTAI, CHINA" : "YANTAI",
                            "DARIEN, CHINA (DALIAN)" : "DALIAN",
                            "FUZHOU, CHINA" : "FUZHOU",
                            "HSINKANG, CHINA (XINGANG)" : "XINGANG",
                            "HSINKANG, CHINA (XINGANG), CN, EXTREME NORTHERN" : "XINGANG",
                            "HUANGPU, CHINA, CN, SOUTHERN" : "HUANGPU",
                            "HUANGPU/WHAMPOA, CHINA" : "HUANGPU",
                            "NANJING, CHINA" : "NANJING",
                            "NANSHA PORT (GUANGZHOU), CHINA" : "NANSHA",
                            "NING BO/NINGPO, CHINA" : "NINGBO",
                            "NING BO/NINGPO, CHINA, CN, NORTHERN" : "NINGBO",
                            "SHANGHAI, CHINA" : "SHANGHAI",
                            "SHANGHAI, CHINA, CN, NORTHERN" : "SHANGHAI",
                            "SHENZHEN, CHINA" : "SHENZHEN",
                            "TSINGTAO, CHINA  (QINGDAO)" : "QINGDAO",
                            "TSINGTAO, CHINA  (QINGDAO), CN, EXTREME NORTHERN" : "QINGDAO",
                            "XIAMEN (AMOU), CHINA" : "XIAMEN",
                            "XIAMEN (AMOU), CHINA, CN, MID SOUTH" : "XIAMEN",
                            "XINHUI, CHINA" : "XINHUI",
                            "YANTIAN, CHINA" : "YANTIAN",
                            "YANTIAN, CHINA, CN, SOUTHERN" : "YANTIAN",
                            "ZHANGJIAGANG, CHINA" : "ZHANGJIAGANG",
                            "ZHENJIANG, CHINA" : "ZHENJIANG",
                            "ZHONGSHAN, CHINA" : "ZHONGSHAN",
                            "ZHONGSHAN, CHINA, CN, SOUTHERN" : "ZHONGSHAN"}
    @staticmethod
    def originConversion(city):
        city = str(city)
        if city in HomeDepotData.originConversionDict: return HomeDepotData.originConversionDict[city]
        else: return city.split(",")[0].upper()



    
    #For "Destination" field
    #Converts from the format in the PO data files to the format in the schedule data file
    #If a city cannot be found in the list below, the city name is extracted and changed to upper case
    #Example: "Dallas, TX" would be changed to "DALLAS"    
    destinationConversionDict = {"ALL OTHER CA. PAC PORTS" : "CALGARY",
                                 "CHICAGO, IL" : "CHICAGO",
                                 "CLEVELAND, OH" : "CLEVELAND",
                                 "DALLAS-FORT WORTH, TX" : "DALLAS",
                                 "FRONT ROYAL, VA" : "FRONT ROYAL",
                                 "HOUSTON, TX" : "HOUSTON",
                                 "LONG BEACH, CA" : "LAX/LGB",
                                 "LOS ANGELES, CA" : "LAX/LGB",
                                 "NEW YORK, NY" : "NYC/NWK",
                                 "NEWARK, NJ" : "NYC/NWK",
                                 "OAKLAND, CA" : "OAKLAND",
                                 "SAVANNAH, GA" : "CHS/SAV",
                                 "SEATTLE, WASHINGTON" : "SEA/TAC",
                                 "ST. LOUIS, MO" : "ST LOUIS",
                                 "TACOMA, WA" : "SEA/TAC",
                                 "TORONTO, ONT, CA." : "TORONTO",
                                 "VANCOUVER, BC, CA." : "VANCOUVER",
                                 "THD BAYTOWN, TX DC" : "THD BAYTOWN, TX DC"}
    @staticmethod
    def destinationConversion(city):
        city = str(city)
        if city in HomeDepotData.destinationConversionDict:
            return HomeDepotData.destinationConversionDict[city]
        else: return city.split(",")[0].upper()

                                 

    def __init__(self, POfn, CSfn, CRfn, RGfn, MQCfn, contractWK):


        #Constants for the data indices in the input files
        #Edit these numbers if the data format is changing
        #Don't forget that data is zero-indexed
        PO_ORIGIN_INDEX = 5
        PO_DESTINATION_INDEX = 9
        PO_EQUIPMENT_INDEX = 7
        PO_SUPPLIER_INDEX = 6
        PO_ESD_INDEX = 2
        PO_FEU_INDEX = None #If this doesn't exist, set it to None

        SCHEDULE_CARRIER_INDEX = 0
        SCHEDULE_ORIGIN_INDEX = 1
        SCHEDULE_DESTINATION_INDEX = 3
        SCHEDULE_CYCLOSE_INDEX = 4
        SCHEDULE_ETD_INDEX = 5
        SCHEDULE_TTIME_INDEX = 6
        SCHEDULE_STRING_INDEX = 7
        
        RATE_CARRIER_INDEX = 0
        RATE_ORIGIN_INDEX = 2
        RATE_DESTINATION_INDEX = 3
        RATE_EQUIPMENT_INDEX = 5
        RATE_COST_INDEX = 6

        ROUTING_ORIGIN_INDEX = 1
        ROUTING_DESTINATION_INDEX = 2
        ROUTING_CSCL_INDEX = 3
        ROUTING_CMDU_INDEX = 4
        ROUTING_EGLV_INDEX = 5
        ROUTING_HJSC_INDEX = 6
        ROUTING_HLCU_INDEX = 7
        ROUTING_MAEU_INDEX = 8
        ROUTING_MOLU_INDEX = 9
        ROUTING_NYKS_INDEX = 10
        ROUTING_OOLU_INDEX = 11

        MQC_CALENDARWEEK_INDEX = 1
        

        ##This is all of the model data
        ##The rest of this method is to create these
        self.I = set() ##Set of all PO types
        self.J = set() ##Set of all assignments
        self.K = set() ##Set of all carriers
        self.L = set() ##Set of all lanes
        self.rows = {} ##Dictionary that takes a PO type as a key and returns a set of excel sheet rows with pos that match. Created for solution extraction after model is completed
        self.count = {} ##Dictionary that takes a PO type as key and returns how many containers of that type are being sent
        self.volume = {} ##Dictionary that takes a lane as a key and returns how much volume is being sent on that lane
        self.R = {} ##Dictionary that takes PO as a key and returns the set of possible assignments that it can use
        self.P = {} ##Dictionary that takes an assignment as a key and returns the set of possible POs that could be sent on it
        self.C = {} ##Dictionary that takes a carrier name as a key and returns the set of assignments that are associated with the carrier
        self.O = {} ##Dictionary that takes a lane as a key and returns the set of POs that are being sent on that lane
        self.MQC = {} ##Dictionary that takes a carrier name as a key and returns the MQC of that carrier
        self.target = {} ##Dictionary that takes a carrier, lane tuple as a key and returns the target for the carrier, lane pairing
        self.cost = {} ##Dictionary that takes a po, assignment pair and returns the cost of assigning the po
        self.MQCpen = {} ##Dictionary that takes a carrier as a key and returns the MQC penalty cost for the carrier
        self.targetpen = {} ##Dictionary that takes a carrier, lane tuple as a key and returns the target penalty for the carrier, lane pairing
        self.unallocated = set() ##Set of POs that cannot be sent
        self.CarrierCaps = {} ##Dictionary to set an FEU capacity restriction for a given lane
        self.alreadyAllocated = dict() #Dictionary of dictionaries that holds information relevant for reallocations
        self.carrierO = {} #Dictionary used to create the origin drop-down menu in the lane preferences on the UI
        self.carrierOD = {} #Dictionary used to create the destination drop-down menu in the lane preferences on the UI
  

        ##Here are the input files for POs and Schedule
        ##UI will populate these fields
        POwb = load_workbook(r"%s" % POfn, data_only=True)
        POws = POwb.active
        Assignwb = load_workbook(r"%s" % CSfn, data_only=True)
        Assignws = Assignwb.active
        Routingwb = load_workbook(r"%s" % RGfn, data_only=True)
        Routingws = Routingwb.active
        Ratewb = load_workbook(r"%s" % CRfn, data_only=True)
        Ratews = Ratewb.active
        MQCwb = load_workbook(r"%s" % MQCfn, data_only=True)
        MQCws = MQCwb.get_sheet_by_name('FEU Capacity')


        self.header = tuple([str(item.value) for item in POws.rows[0]])

        ##Populates I, L, O rows, count, and volume
        for row in POws.rows[1:]:
            poRow = row
            po = PO()
            try:
                po.ESD = str(1 + row[PO_ESD_INDEX].value.weekday())
                po.ESDDate = row[PO_ESD_INDEX].value
                po.Supplier = str(row[PO_SUPPLIER_INDEX].value).replace(":", "")
                po.Origin = HomeDepotData.originConversion(row[PO_ORIGIN_INDEX].value)
                po.Destination = HomeDepotData.destinationConversion(row[PO_DESTINATION_INDEX].value)
                po.Equipment = HomeDepotData.equipmentConversion(row[PO_EQUIPMENT_INDEX].value)
            except:
                print("There appears to have been missing data. The following row was ignored: ")
                print(row)
                continue
            if PO_FEU_INDEX is None: po.FEU = HomeDepotData.feuConversion(po.Equipment)
            else: po.FEU = int(row[PO_FEU_INDEX].value)

            newRow = tuple([item.value for item in poRow])
            
            if po in self.rows:
                self.rows[po].add(newRow)
                self.count[po] += 1
            else:
                self.I.add(po)
                self.rows[po] = set()
                self.rows[po].add(newRow)
                self.count[po] = 1
            lane = (po.Origin, po.Destination)
            if lane in self.O:
                self.O[lane].add(po)
                self.volume[lane] += po.FEU
            else:
                self.O[lane] = set()
                self.O[lane].add(po)
                self.L.add(lane)
                self.volume[lane] = po.FEU


        ##Populates J and K
        for row in Assignws.rows[1:]:
            assignment = Assignment()
            assignment.Carrier = str(row[SCHEDULE_CARRIER_INDEX].value)
            assignment.Origin = HomeDepotData.originConversion(row[SCHEDULE_ORIGIN_INDEX].value)
            assignment.Destination = HomeDepotData.destinationConversion(row[SCHEDULE_DESTINATION_INDEX].value)
            assignment.CYClose = str(row[SCHEDULE_CYCLOSE_INDEX].value)
            assignment.ETD = str(row[SCHEDULE_ETD_INDEX].value)
            assignment.TTime = str(row[SCHEDULE_TTIME_INDEX].value)
            assignment.String = str(row[SCHEDULE_STRING_INDEX].value)
            self.K.add(assignment.Carrier)
            self.J.add(assignment)

        ##Populates R
        for po in self.I:
            self.R[po] = set()
            for assignment in self.J:
                if HomeDepotData.cityEqual(po.Origin, assignment.Origin) and HomeDepotData.cityEqual(po.Destination, assignment.Destination):
                    self.R[po].add(assignment)

        ##Populates P
        for assignment in self.J:
            self.P[assignment] = set()
            for po in self.I:
                if HomeDepotData.cityEqual(po.Origin, assignment.Origin) and HomeDepotData.cityEqual(po.Destination, assignment.Destination):
                    self.P[assignment].add(po)

        ##Populates C, MQC
        MQCrow = MQCws.rows[contractWK+2]
        
        #The carrierCol dictionary contains the column index for each carrier in the MQC Report file
        #This will need to be changed if the format/carriers change
        carrierCol = {"CMDU":5, "CSCL":6, "EGLV":7, "HJSC":8, "HLCU":9, "MAEU":10, "MOLU":11, "NYKS":12, "OOLU":13}
        carrierMQCPen = {"CMDU":10000 , "CSCL":10000, "EGLV":10000, "HJSC":10000, "HLCU":10000, "MAEU":10000, "MOLU":10000, "NYKS":10000, "OOLU":10000}
        
        for carrier in self.K:
            self.C[carrier] = set()
            self.MQC[carrier] = int(MQCrow[carrierCol[carrier]].value)
            
            self.MQCpen[carrier] = carrierMQCPen[carrier] 
        for assignment in self.J:
            self.C[assignment.Carrier].add(assignment)


        ##Populates carrierO and carrierOD
        #Used for drop down menus in UI
        for c in self.C:
            s = set()
            d = set()
            for o in self.C[c]:
                s.add(o.Origin)
                d.add(o.Destination)
            self.carrierO[c] = s
            self.carrierOD[c] = d


        ##This code block populates cost
        costDict = {}
        rows = Ratews.rows[1:]
        total = 0
        count = 0
        for row in rows:
            key = (str(row[RATE_CARRIER_INDEX].value).upper(), City(str(row[RATE_ORIGIN_INDEX].value).upper()), City(str(row[RATE_DESTINATION_INDEX].value).upper()), str(row[RATE_EQUIPMENT_INDEX].value).upper())
            costDict[key] = int(row[RATE_COST_INDEX].value)
            total += int(row[RATE_COST_INDEX].value)
            count += 1
        avgCost = int(total/count)

        for i in self.I:
            for j in self.R[i]:
                try:
                    equipment = i.Equipment
                    if equipment == "CFS" or equipment == "NONE":
                        equipment = "40HIGH"
                        self.cost[i, j] = i.FEU * costDict[(str(j.Carrier).upper(), City(str(j.Origin).upper()), City(str(j.Destination).upper()), equipment)]
                    else:
                        self.cost[i, j] = costDict[(str(j.Carrier).upper(), City(str(j.Origin).upper()), City(str(j.Destination).upper()), equipment)]
                except:
                    self.cost[i, j] = avgCost #If cost cannot be found, uses the overall average as a placeholder

                
        ##Populates unallocated, and removes infeasible POs from set I
        for i in self.I:
            if len(self.R[i]) == 0:
                self.unallocated.add(i)
        for i in self.unallocated:
            self.I.remove(i)


        ##Populates target, targetpen
        for k in self.K:
            for l in self.L:
                self.target[k, l] = 0
                key = (k, City(l[0].upper()), City(l[1].upper()), "40HIGH")
                if key in costDict:
                    self.targetpen[k,l] = costDict[(k, City(l[0].upper()), City(l[1].upper()), "40HIGH")]
                else:
                    self.targetpen[k,l] = 4341 #Placeholder - average 40HIGH cost
                
        carriers = {0:"CSCL", 1:"CMDU", 2:"EGLV", 3:"HJSC", 4:"HLCU", 5:"MAEU", 6:"MOLU", 7:"NYKS", 8:"OOLU"}
        i = 0
        for row in Routingws.rows[1:]:
            origin = str(row[1].value).split(",")[0].upper()
            destination = str(row[2].value).split(",")[0].upper()
            l = origin, destination
            if l in self.L:
                for i in range(9):
                    k = carriers[i]
                    self.target[k,l] = float(self.volume[l])*float(row[i+3].value)


    ##Function to pull in a list of carrier/lane pairs created within the UI and removes the strings from the previously defined data structures
    def removeStrings(self, stringsRemoved):
        List = []
        for p in self.P:
            string = (p.Carrier, p.Origin, p.Destination)
            if string in stringsRemoved:
                List.append(p)
        for a in List:
            del self.P[a]
         
        List = []
        for k in self.K:
            for a in self.C[k]:
                lane = (a.Carrier, a.Origin, a.Destination)
                if lane in stringsRemoved:
                    List.append((k,a))
        for a in List:
            self.C[a[0]].remove(a[1])

        List = []
        for k in self.R:
            for a in self.R[k]:
                lane = (a.Carrier, a.Origin, a.Destination)
                if lane in stringsRemoved:
                    List.append((k, a))
        for a in List:
            self.R[a[0]].remove(a[1])

        List = []
        for a in stringsRemoved:
            tup = (a[0], (a[1], a[2]))
            if tup in self.target:
                del self.target[tup]


    ##Functions to add in and remove carrier/lane capacity restrictions
    def addCapacity(self, carrier, lane, capacity):
        self.CarrierCaps[carrier, lane] = capacity
        print('These carrier/lanes will be capped: ', self.CarrierCaps)
        return None

    def removeCapacity(self, carrier, lane, capacity):
        del self.CarrierCaps[carrier, lane]
        print('These carrier/lanes will be capped: ', self.CarrierCaps)
        return None
        

    @staticmethod
    def cityEqual(city1, city2):
        city1 = city1.split(",")[0].upper()
        city2 = city2.split(",")[0].upper()
        if city1 == city2:
            return True
        elif city1 in City.abbreviations:
            return city2 in City.abbreviations[city1]
        elif city2 in City.abbreviations:
            return city1 in City.abbreviations[city2]
        elif city1 in City.citiesToAbbr and City.citiesToAbbr[city1] in City.abbreviations:
            return city2 in City.abbreviations[City.citiesToAbbr[city1]]
        elif city2 in City.citiesToAbbr and City.citiesToAbbr[city2] in City.abbreviations:
            return city1 in City.abbreviations[City.citiesToAbbr[city2]]
        
        return False

    ##Function to cross analyze allocation with POs and re-allocate the new and late-cut POs only
    ##Creates new data structures to be ran through the model
    def getOriginalAllocation(self, OAfn):
        OUTPUT_ORIGIN_INDEX = 5
        OUTPUT_DESTINATION_INDEX = 9
        OUTPUT_EQUIPMENT_INDEX = 7
        OUTPUT_SUPPLIER_INDEX = 6
        OUTPUT_ESD_INDEX = 2
        OUTPUT_CARRIER_INDEX = 12
        OUTPUT_ETD_INDEX = 14
        OUTPUT_ETA_INDEX = 15
        OUTPUT_STRING_INDEX = 16
        
        OAwb = load_workbook(r"%s" % OAfn, data_only=True)
        OAws = OAwb.active
        for row in OAws.rows[1:]:
            assignment = Assignment()
            assignment.Carrier = str(row[OUTPUT_CARRIER_INDEX].value)
            if assignment.Carrier == "UNALLOCATED": continue
            assignment.Origin = HomeDepotData.originConversion(row[OUTPUT_ORIGIN_INDEX].value)
            assignment.Destination = HomeDepotData.destinationConversion(row[OUTPUT_DESTINATION_INDEX].value)
            ETD = datetime.strptime(row[OUTPUT_ETD_INDEX].value, "%m/%d/%y")
            ETA = datetime.strptime(row[OUTPUT_ETA_INDEX].value, "%m/%d/%y")
            assignment.ETD = str(1 + ETD.weekday())
            assignment.TTime = str((ETA-ETD).days)
            assignment.String = str(row[OUTPUT_STRING_INDEX].value)
            for other in self.J:
                if assignment.similar(other):
                    assignment.CYClose = other.CYClose
                    break
            ESD = datetime.strptime(row[OUTPUT_ESD_INDEX].value, "%m/%d/%y")

            po = PO()
            po.ESD = str(1 + ESD.weekday())
            po.ESDDate = ESD
            po.Supplier = str(row[OUTPUT_SUPPLIER_INDEX].value).replace(":", "")
            po.Origin = HomeDepotData.originConversion(row[OUTPUT_ORIGIN_INDEX].value)
            po.Destination = HomeDepotData.destinationConversion(row[OUTPUT_DESTINATION_INDEX].value)
            po.Equipment = HomeDepotData.equipmentConversion(row[OUTPUT_EQUIPMENT_INDEX].value)
            po.FEU = HomeDepotData.feuConversion(po.Equipment)

            if po in self.alreadyAllocated:
                if assignment in self.alreadyAllocated[po]:
                    self.alreadyAllocated[po][assignment] += 1
                else:
                    self.alreadyAllocated[po][assignment] = 1
            else:
                self.alreadyAllocated[po] = dict()
                self.alreadyAllocated[po][assignment] = 1


#Dummy class purely for comparison purposes       
class City:
    
    abbreviations = {"BOM/NSA" : ["BOMBAY", "NHAVA SHIVA"],
                 "FXT/THM/TIL/STH" : ["FELIXTOWE", "SOUTHAMPTON", "THAMESPORT", "TILBURY"],
                 "ITJ/SFS/NAV" : ["ITAJAI", "NAVAGANTES"],
                 "LCH/BKK" : ["BANGKOK", "LAEM CHABANG"],
                 "SZX/SHK/CWN/YTN" : ["CHIWAN", "SHEKOU", "SHENZHEN", "YANTIAN"],
                 "XIN/TJN" : ["XINGANG", "TIANJIN"],
                 "CHS/SAV" : ["CHARLESTON", "SAVANNAH"],
                 "SAV/CHS" : ["CHARLESTON", "SAVANNAH"],
                 "LAX/LGB" : ["LONG BEACH", "LOS ANGELES"],
                 "NY/NJ" : ["NEWARK", "NEW YORK"],
                 "NYC/NWK" : ["NEWARK", "NEW YORK"],
                 "SEA/TAC" : ["SEATTLE", "TACOMA"]}

    citiesToAbbr = {"BOMBAY" : "BOM/NSA",
                    "NHAVA SHIVA" : "BOM/NSA",
                    "FELIXTOWE" : "FXT/THM/TIL/STH",
                    "SOUTHAMPTON" : "FXT/THM/TIL/STH",
                    "THAMESPORT" : "FXT/THM/TIL/STH",
                    "TILBURY" : "FXT/THM/TIL/STH",
                    "ITAJAI" : "ITJ/SFS/NAV",
                    "NAVAGANTES" : "ITJ/SFS/NAV",
                    "BANGKOK" : "LCH/BKK",
                    "LAEM CHABANG" : "LCH/BKK",
                    "CHIWAN" : "SZX/SHK/CWN/YTN",
                    "SHEKOU" : "SZX/SHK/CWN/YTN",
                    "SHENZHEN" : "SZX/SHK/CWN/YTN",
                    "YANTIAN" : "SZX/SHK/CWN/YTN",
                    "XINGANG" : "XIN/TJN",
                    "TIANJIN" : "XIN/TJN",
                    "CHARLESTON" : "CHS/SAV",
                    "SAVANNAH" : "CHS/SAV",
                    "LONG BEACH" : "LAX/LGB",
                    "LOS ANGELES" : "LAX/LGB",
                    "NEWARK" : "NYC/NWK",
                    "NEW YORK" : "NYC/NWK",
                    "SEATTLE" : "SEA/TAC",
                    "TACOMA" : "SEA/TAC"}

    
    def __init__(self, city):
        if city in City.citiesToAbbr: self.city = City.citiesToAbbr[city]
        else: self.city = city

    def __hash__(self):
        if self.city not in City.citiesToAbbr: return hash(self.city)
        else: return hash(City.citiesToAbbr[self.city])

    def __eq__(self, other):
        return HomeDepotData.cityEqual(self.city, other.city) 
        

        
                     
    
