from openpyxl import Workbook
from datetime import *

def outputModelResults(model):
    wb = Workbook(write_only = True)
    ws = wb.create_sheet()
    pivotws = wb.create_sheet()
    
    header = list(model.data.header)
    for item in ["Allocated Carrier", "Vessel Name", "Vessel ETD", "Vessel ETA", "String"]:
        header.append(item)
    ws.append(header)

    header = ["Origin", "Destination", "CSCL", "CMDU", 'EGLV', 'HJSC', 'HLCU', 'MAEU', 'MOLU', 'NYKS', 'OOLU', 'FEU COUNT']
    pivotws.append(header)

    totalCost = 0
    MQCError = sum(model.y[item].X for item in model.y)
    totalVolume = sum(item.FEU*model.data.count[item] for item in model.data.I)
    targetError = sum(model.z[item].X for item in model.z)
    
    for l in model.data.L:
        laneTotal=0
        row = [l[0], l[1]]
        for k in header[2:11]:
            laneTotal += (sum(sum(model.x[i,j].X*i.FEU for j in model.data.C[k].intersection(model.data.R[i])) for i in model.data.O[l]))
        for k in header[2:11]:
            if laneTotal != 0:
                allocationPercent = str(round(100*(sum(sum(model.x[i,j].X*i.FEU for j in model.data.C[k].intersection(model.data.R[i])) for i in model.data.O[l]))/laneTotal, 2))
                row.append(allocationPercent)
        row.append(laneTotal)
        pivotws.append(row)
    
    for item in model.x:
        var = model.x[item]
        po = item[0]
        assignment = item[1]
        count = int(var.X)
        rate = model.data.cost[po, assignment]
        cost = rate*count
        row = []

        
        for i in range(count):
            row = []
            try:
                for cell in list(model.data.rows[po].pop()):
                    if type(cell) == datetime:
                        row.append(cell.strftime("%m/%d/%y"))
                    else:
                        row.append(cell)

            except:
                print("Duplicate entry found and discarded")
                
            row.append(assignment.Carrier)
            row.append(assignment.Vessel)

            daysToCYClose = int(assignment.CYClose) - int(po.ESD)
            daysToCYClose = daysToCYClose if daysToCYClose >= 0 else (daysToCYClose + 7)
            daysToETD = int(assignment.ETD) - int(assignment.CYClose)
            daysToETD = daysToETD if daysToETD >=0 else (daysToETD + 7)
            ETD = po.ESDDate + timedelta(days = daysToCYClose + daysToETD)
            row.append(ETD.strftime("%m/%d/%y"))

            ETA = ETD + timedelta(days = int(assignment.TTime))
            row.append(ETA.strftime("%m/%d/%y"))

            totalCost += rate
            
            row.append(assignment.String)
            ws.append(row)
    
    for item in model.data.unallocated:
        while model.data.rows[item]:
            row = list(model.data.rows[item].pop())
            row.append("UNALLOCATED")
            ws.append(row)

    print("Cost: " + '${:,.2f}'.format(totalCost))
    print("Weighted Average % Allocation Error: " + str(100*targetError/totalVolume))
    print("Weighted Average MQC Error: " + str(100*MQCError/totalVolume))


    wb.save('output.xlsx')
